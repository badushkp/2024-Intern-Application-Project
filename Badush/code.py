import pdfplumber
import re
from openpyxl import Workbook
from openpyxl.styles import Alignment

# Define input PDF files
pdf_files = ["iee.pdf"]

# Initialize workbook and worksheet
wb = Workbook()
ws = wb.active

# Define main heading and subheadings
main_heading = "Basic Descriptors"
subheadings = ["Title", "Year", "Author", "Journal", "Type(eg..review,theory)"]

# Append subheadings under the main heading
ws.append([main_heading])
ws.append(subheadings)

# Function to process PDF and extract information
def process_pdf(pdf_file):
    data = []  # To store extracted data from all pages
    with pdfplumber.open(pdf_file) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            extracted_info = extract_info_from_text(text)
            if extracted_info:
                data.append(extracted_info)
    return data

# Function to extract information from text
def extract_info_from_text(text):
    # Define patterns for title, year, author, journal, and type
    title_pattern = r'Title:(.*?)\n'
    year_pattern = r'Year:(\d{4})\n'
    author_pattern = r'Author:(.*?)\n'
    journal_pattern = r'Journal:(.*?)\n'
    type_pattern = r'Type:(.*?)\n'

    # Extract information using regular expressions
    title = re.search(title_pattern, text, re.IGNORECASE)
    year = re.search(year_pattern, text)
    author = re.search(author_pattern, text, re.IGNORECASE)
    journal = re.search(journal_pattern, text, re.IGNORECASE)
    type_match = re.search(type_pattern, text, re.IGNORECASE)

    # Extracted information
    extracted_info = [
        title.group(1).strip() if title else None,
        year.group(1).strip() if year else None,
        author.group(1).strip() if author else None,
        journal.group(1).strip() if journal else None,
        type_match.group(1).strip() if type_match else None
    ]

    return extracted_info

# Process each PDF file in the list
for pdf_file in pdf_files:
    data = process_pdf(pdf_file)
    for info in data:
        ws.append(info)

# Merge cells for the main heading and set alignment
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(subheadings))
main_heading_cell = ws.cell(row=1, column=1)
main_heading_cell.alignment = Alignment(horizontal='center')

# Save the workbook to an Excel file
excel_file_path = "ext.xlsx"
wb.save(excel_file_path)

print(f"Data has been saved to '{excel_file_path}'")